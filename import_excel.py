import requests
import json
import time
from datetime import datetime
import openpyxl

# API Configuration
API_URL = "https://wki-wip-api.onrender.com/api"  # Production

print("🚀 Starting Excel Import to MongoDB via API...")
print(f"📡 Target: {API_URL}\n")

# Read Excel file
excel_file = 'current_wip.xlsx'  # Change this to your Excel file name

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    print(f"📊 Workbook loaded: {excel_file}")
    print(f"📋 Sheets found: {workbook.sheetnames}\n")
except FileNotFoundError:
    print(f"❌ File not found: {excel_file}")
    print("\nPlease export your Google Sheets as Excel (.xlsx):")
    print("  1. In Google Sheets, go to File → Download → Microsoft Excel (.xlsx)")
    print("  2. Save it in this folder as 'current_wip.xlsx'")
    print("  3. Run this script again")
    exit(1)

# Process each sheet
for sheet_name in workbook.sheetnames:
    print(f"\n{'='*80}")
    print(f"Processing sheet: {sheet_name}")
    print(f"{'='*80}\n")
    
    sheet = workbook[sheet_name]
    
    # Get header row
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f"Headers: {', '.join(headers[:5])}...\n")
    
    # Parse rows
    orders = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not any(row[:3]):  # Check first 3 columns
            continue
        
        # Create order dictionary
        order_dict = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else None
            order_dict[header] = str(value).strip() if value else ''
        
        # Map to database schema
        order = {
            'customer': order_dict.get('Customer', '').strip(),
            'unit': order_dict.get('UNIT', '').strip(),
            'ro': order_dict.get('R.O.', '').strip(),
            'bay': order_dict.get('Bay #', '').strip(),
            'firstShift': order_dict.get('First Shift Notes', '').strip(),
            'secondShift': order_dict.get('Second Shift Notes', '').strip(),
            'orderedParts': order_dict.get('Ordered parts/ETA and TCS case #s', '').strip(),
            'triageNotes': order_dict.get('Triage Notes ', '').strip(),  # Note the space
            'quoteStatus': order_dict.get('Quote Status', '').strip(),
            'repairCondition': order_dict.get('Repair Condition', '').strip(),
            'contactInfo': order_dict.get('Contact Info', '').strip(),
            'accountStatus': order_dict.get('Account Status', '').strip(),
            'customerStatus': order_dict.get('Customer Status', '').strip(),
            'call': order_dict.get('Call', '').strip(),
        }
        
        # Only add if required fields are present
        if order['customer'] and order['unit'] and order['ro']:
            orders.append(order)
    
    if not orders:
        print(f"⚠️  No valid orders found in sheet '{sheet_name}'\n")
        continue
    
    print(f"📊 Found {len(orders)} valid orders in '{sheet_name}'\n")
    
    # Determine if this is archived data or current WIP
    sheet_lower = sheet_name.lower()
    is_current = 'current' in sheet_lower or 'wip' in sheet_lower
    
    if is_current:
        # Import to current orders
        print(f"📥 Importing to CURRENT orders...\n")
        endpoint = f"{API_URL}/orders"
    else:
        # This is an archive sheet - determine the month
        archive_month = sheet_name  # e.g., "December 2025"
        print(f"📦 Importing to ARCHIVE: {archive_month}\n")
        
        # We'll need to:
        # 1. Create the order in current
        # 2. Immediately archive it to the specific month
        # Or just add dateAdded to match the archive period
    
    success_count = 0
    error_count = 0
    
    print("=" * 100)
    print(f"{'#':<4} {'Customer':<35} {'Unit':<10} {'RO':<10} {'Status':<10}")
    print("=" * 100)
    
    for idx, order in enumerate(orders, 1):
        try:
            if is_current:
                # Add current date
                order['dateAdded'] = datetime.now().strftime('%Y-%m-%d')
                response = requests.post(endpoint, json=order, timeout=10)
            else:
                # For archived sheets, we need to archive them
                # First create as regular order
                order['dateAdded'] = datetime.now().strftime('%Y-%m-%d')
                response = requests.post(f"{API_URL}/orders", json=order, timeout=10)
                
                if response.status_code in [200, 201]:
                    # Then immediately archive it
                    order_id = response.json().get('_id')
                    archive_response = requests.post(
                        f"{API_URL}/orders/{order_id}/archive",
                        json={'archiveMonth': archive_month},
                        timeout=10
                    )
                    response = archive_response  # Use archive response for status
            
            if response.status_code in [200, 201]:
                success_count += 1
                status = f"✅ {response.status_code}"
            else:
                error_count += 1
                status = f"❌ {response.status_code}"
            
        except Exception as e:
            error_count += 1
            status = f"❌ Error"
        
        customer = order['customer'][:33]
        unit = order['unit'][:8]
        ro = order['ro'][:8]
        print(f"{idx:<4} {customer:<35} {unit:<10} {ro:<10} {status:<10}")
        
        time.sleep(0.1)  # Small delay
    
    print("=" * 100)
    print(f"\n{'='*80}")
    print(f"Sheet '{sheet_name}' Summary:")
    print(f"   ✅ Success: {success_count}")
    print(f"   ❌ Errors: {error_count}")
    print(f"   📈 Total: {len(orders)}")
    print(f"{'='*80}\n")

print("\n" + "="*80)
print("🎉 EXCEL IMPORT COMPLETE!")
print("="*80)
